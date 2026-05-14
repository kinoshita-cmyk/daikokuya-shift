"""
シフト表 Excel 出力（テンプレート完全互換版・A4縦）
================================================
data/may_2026_shift.xlsx の「シフト表2026年5月」シートと完全に同一の
ページ設定・列幅・行高さ・フォントで書き出す。

元テンプレートの仕様:
- A4 縦（portrait）、1ページに収めて印刷、左右余白 0
- フォント: ＭＳ Ｐゴシック
- 行高: タイトル90pt、コメント90pt、凡例45pt、ヘッダー45pt、データ45pt
- フォントサイズ: タイトル60pt、コメント43pt、凡例36pt、ヘッダー24pt、データ35pt、注意書き27pt
- 列幅: A4縦1ページに収まるよう、従業員列と人員少欄を調整
"""

from __future__ import annotations
import warnings
from datetime import date
from pathlib import Path
from typing import Optional
from calendar import monthrange

import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import MonthlyShift, Store, OperationMode
from .paths import OUTPUT_DIR
from .employees import ALL_EMPLOYEES
from .rules import STORE_KEYHOLDERS, SUZURAN_KEY_SUPPORT_FROM_OMIYA, get_capacity


# 出力時の従業員列順（運用に慣れた順番、テンプレートと同じ）
EXPORT_COLUMN_ORDER = [
    "山本", "板倉", "今津", "鈴木", "田中", "岩野", "大塚", "南",
    "黒澤", "牧野", "春山", "下地", "大類", "長尾", "野澤", "下田",
    "楯", "土井", "顧問",
]

WEEKDAY_JP = ["月", "火", "水", "木", "金", "土", "日"]

# テンプレートで使用するフォント
JP_FONT = "ＭＳ Ｐゴシック"

# 列幅: 画面表示でも"###"が出にくく、A4縦1ページ印刷時に収まる幅。
COLUMN_WIDTHS = {
    "A": 2.0,
    "B": 8.5,         # 日付（2桁数字でも#が出ないように）
    "C": 7.0,         # 曜日（1文字）
    "D": 11.0,        # 山本
    "E": 13.0,        # 板倉
    "F": 13.0,        # 今津
    "G": 11.5,        # 鈴木
    "H": 13.0,        # 田中
    "I": 11.0,        # 岩野
    "J": 11.5,        # 大塚
    "K": 11.0,        # 南
    "L": 13.0,        # 黒澤
    "M": 13.0,        # 牧野
    "N": 11.0,        # 春山
    "O": 13.0,        # 下地
    "P": 13.0,        # 大類
    "Q": 13.0,        # 長尾
    "R": 13.0,        # 野澤
    "S": 13.0,        # 下田
    "T": 11.0,        # 楯
    "U": 13.0,        # 土井
    "V": 13.0,        # 顧問
    "W": 8.5,         # 日付（右）
    "X": 7.0,         # 曜日（右）
    "Y": 21.0,        # 人員少（店舗別マークが収まる幅）
    "Z": 18.0,        # 鍵確認
}

SHORT_STAFF_STORE_LABELS = {
    Store.AKABANE: "○赤羽",
    Store.HIGASHIGUCHI: "□東口",
    Store.OMIYA: "△大宮",
    Store.NISHIGUCHI: "☆西口",
    Store.SUZURAN: "◆すずらん",
}

DEFAULT_FOOTER_NOTES = [
    "※25日までに翌月のお休み又は出勤希望日を、ご連絡ください。（お忘れなく！！）",
    "※出勤基準日数（の目安）と違いがある場合は、希望するお休み日数と消化する有給休暇日数もお願いします。",
    "※出勤簿は月末までに、赤羽に到着するように提出してください。",
]


def detect_key_warnings_by_store(shift: MonthlyShift) -> dict[int, dict[Store, str]]:
    """鍵担当がいない店舗を日付・店舗別に検出する。"""
    warnings_by_store: dict[int, dict[Store, str]] = {}
    days_in_month = monthrange(shift.year, shift.month)[1]
    for d in range(1, days_in_month + 1):
        mode = shift.operation_modes.get(d, OperationMode.NORMAL)
        if mode == OperationMode.CLOSED:
            continue
        capacity_map = get_capacity(mode)
        weekday = date(shift.year, shift.month, d).weekday()
        day_assignments = shift.get_day_assignments(d)
        for store, keyholders in STORE_KEYHOLDERS.items():
            cap = capacity_map.get(store)
            if cap is None:
                continue
            if weekday in cap.closed_dow:
                continue
            workers = [a.employee for a in day_assignments if a.store == store]
            if not workers:
                continue
            if any(name in keyholders for name in workers):
                continue
            status = "missing"
            if store == Store.SUZURAN:
                omiya_workers = [
                    a.employee for a in day_assignments
                    if a.store == Store.OMIYA
                ]
                if any(name in SUZURAN_KEY_SUPPORT_FROM_OMIYA for name in omiya_workers):
                    status = "support"
            warnings_by_store.setdefault(d, {})[store] = status
    return warnings_by_store


def format_key_warning_text(statuses: dict[Store, str]) -> str:
    """鍵欄に出す短い文字列。"""
    if not statuses:
        return ""
    order = list(SHORT_STAFF_STORE_LABELS)
    labels = []
    for store in sorted(
        statuses,
        key=lambda s: order.index(s) if s in order else len(order),
    ):
        base = SHORT_STAFF_STORE_LABELS.get(store, getattr(store, "value", str(store)))
        prefix = "応援" if statuses[store] == "support" else "鍵"
        labels.append(f"{prefix}{base}")
    return " ".join(labels)


def export_shift_to_excel(
    shift: MonthlyShift,
    output_path,  # str or Path
    title: Optional[str] = None,
    header_comments: Optional[list[str]] = None,
    footer_notes: Optional[list[str]] = None,
    short_staff_days: Optional[object] = None,
    key_warnings_by_store: Optional[dict[int, dict[Store, str]]] = None,
) -> Path:
    """
    シフトを Excel に出力する（テンプレート完全互換・A4縦）。

    Args:
        shift: 出力対象のシフト
        output_path: 出力先パス
        title: タイトル（デフォルト: "YYYY年M月の目標とシフト表  決定版"）
        header_comments: タイトル下の3行コメント（B3, B4, B5に対応）
        footer_notes: 表の下の注意書き（任意の行数）
        short_staff_days: 「人員少」マークを付ける日のリスト、または {日: {店舗}} の辞書
        key_warnings_by_store: 鍵確認の {日: {店舗: "missing"|"support"}}

    Returns:
        実際に書き込んだファイルパス
    """
    warnings.filterwarnings("ignore")
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    short_staff_days = short_staff_days or []
    key_warnings_by_store = key_warnings_by_store or detect_key_warnings_by_store(shift)
    days_in_month = monthrange(shift.year, shift.month)[1]

    if title is None:
        title = f"{shift.year}年{shift.month}月の目標とシフト表  決定版"
    header_comments = list(header_comments or ["", "", ""])[:3]
    while len(header_comments) < 3:
        header_comments.append("")
    footer_notes = footer_notes or DEFAULT_FOOTER_NOTES

    def _short_staff_text(day: int) -> str:
        if isinstance(short_staff_days, dict):
            stores = short_staff_days.get(day, set())
            order = list(SHORT_STAFF_STORE_LABELS)
            labels = []
            for store in sorted(
                stores,
                key=lambda s: order.index(s) if s in order else len(order),
            ):
                labels.append(SHORT_STAFF_STORE_LABELS.get(store, getattr(store, "value", str(store))))
            return " ".join(labels)
        return "△" if day in short_staff_days else ""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"シフト表{shift.year}年{shift.month}月"

    # ============================================================
    # フォント定義（テンプレート完全準拠）
    # ============================================================
    title_font = Font(name=JP_FONT, size=60, bold=True)
    comment_font = Font(name=JP_FONT, size=43, bold=True)
    legend_font = Font(name=JP_FONT, size=36, bold=True)
    header_font = Font(name=JP_FONT, size=24, bold=True)
    cell_font = Font(name=JP_FONT, size=35, bold=False)
    note_font = Font(name=JP_FONT, size=27, bold=True)

    # 罫線・スタイル
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    # shrink_to_fit=True で大きな文字も列幅に収まるように自動縮小（### を防ぐ）
    center = Alignment(horizontal="center", vertical="center", shrink_to_fit=True)
    left_center = Alignment(horizontal="left", vertical="center", wrap_text=False, shrink_to_fit=True)
    short_fill = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
    key_fill = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")

    LAST_COL = 26  # Z列

    def _style_range_border(row: int, start_col: int = 2, end_col: int = LAST_COL) -> None:
        """結合行でも外枠と内部罫線が印刷で崩れないよう、範囲全体に罫線を入れる。"""
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = border

    # ============================================================
    # B2:Z2 タイトル
    # ============================================================
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=LAST_COL)
    c = ws.cell(row=2, column=2, value=title)
    c.font = title_font
    c.alignment = center
    _style_range_border(2)

    # ============================================================
    # B3:Z3, B4:Z4, B5:Z5 コメント欄（自由記述）
    # ============================================================
    for i in range(3):
        row = 3 + i
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=LAST_COL)
        c = ws.cell(row=row, column=2, value=header_comments[i])
        c.font = comment_font
        c.alignment = center
        _style_range_border(row)

    # ============================================================
    # B6:Z6 凡例
    # ============================================================
    legend = f"{shift.year}年{shift.month}月のシフト表　○赤羽　□東口　△大宮　☆西口　◆すずらん"
    ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=LAST_COL)
    c = ws.cell(row=6, column=2, value=legend)
    c.font = legend_font
    c.alignment = center
    _style_range_border(6)

    # ============================================================
    # 7行目 ヘッダー（左日付・従業員名・右日付・人員少）
    # ※結合セルは結合範囲内の全セルに罫線を適用しないと辺が描画されないので注意
    # ============================================================
    # B7:C7 結合（月ラベル左）
    ws.merge_cells(start_row=7, start_column=2, end_row=7, end_column=3)
    c = ws.cell(row=7, column=2, value=f"{shift.month}月")
    c.font = header_font
    c.alignment = center
    for col in range(2, 4):
        ws.cell(row=7, column=col).border = border

    # D7-V7: 従業員名
    for i, name in enumerate(EXPORT_COLUMN_ORDER):
        col = 4 + i
        c = ws.cell(row=7, column=col, value=name)
        c.font = header_font
        c.alignment = center
        c.border = border

    # W7:X7 結合（月ラベル右）
    ws.merge_cells(start_row=7, start_column=23, end_row=7, end_column=24)
    c = ws.cell(row=7, column=23, value=f"{shift.month}月")
    c.font = header_font
    c.alignment = center
    for col in range(23, 25):
        ws.cell(row=7, column=col).border = border

    # Y7: 人員少
    c = ws.cell(row=7, column=25, value="人員少")
    c.font = header_font
    c.alignment = center
    c.border = border

    # Z7: 鍵
    c = ws.cell(row=7, column=26, value="鍵")
    c.font = header_font
    c.alignment = center
    c.border = border

    # ============================================================
    # 8行目以降 データ
    # ============================================================
    for d in range(1, days_in_month + 1):
        row = 7 + d
        weekday = date(shift.year, shift.month, d).weekday()
        wd = WEEKDAY_JP[weekday]

        # B列: 日付（左）
        c = ws.cell(row=row, column=2, value=d)
        c.font = cell_font
        c.alignment = center
        c.border = border

        # Z列: 鍵確認
        key_text = format_key_warning_text(key_warnings_by_store.get(d, {}))
        if key_text:
            c = ws.cell(row=row, column=26, value=key_text)
            c.fill = key_fill
        else:
            c = ws.cell(row=row, column=26, value="")
        c.font = cell_font
        c.alignment = center
        c.border = border

        # C列: 曜日（左）
        c = ws.cell(row=row, column=3, value=wd)
        c.font = cell_font
        c.alignment = center
        c.border = border

        # D列〜V列: 従業員配属
        for i, name in enumerate(EXPORT_COLUMN_ORDER):
            col = 4 + i
            a = shift.get_assignment(name, d)
            value = a.store.value if a else ""
            c = ws.cell(row=row, column=col, value=value)
            c.font = cell_font
            c.alignment = center
            c.border = border

        # W列: 日付（右）
        c = ws.cell(row=row, column=23, value=d)
        c.font = cell_font
        c.alignment = center
        c.border = border

        # X列: 曜日（右）
        c = ws.cell(row=row, column=24, value=wd)
        c.font = cell_font
        c.alignment = center
        c.border = border

        # Y列: 人員少
        short_text = _short_staff_text(d)
        if short_text:
            c = ws.cell(row=row, column=25, value=short_text)
            c.fill = short_fill
        else:
            c = ws.cell(row=row, column=25, value="")
        c.font = cell_font
        c.alignment = center
        c.border = border

    # ============================================================
    # 末尾: 注意書き（行間1空けて）
    # ============================================================
    end_row = 8 + days_in_month + 1  # データ最終行 + 1
    for i, note in enumerate(footer_notes):
        row = end_row + i
        ws.merge_cells(
            start_row=row, start_column=2, end_row=row, end_column=LAST_COL,
        )
        c = ws.cell(row=row, column=2, value=note)
        c.font = note_font
        c.alignment = left_center

    # ============================================================
    # 列幅（テンプレート完全準拠）
    # ============================================================
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ============================================================
    # 行高さ（テンプレート完全準拠）
    # ============================================================
    ws.row_dimensions[2].height = 90.0   # タイトル
    ws.row_dimensions[3].height = 90.0   # コメント1
    ws.row_dimensions[4].height = 90.0   # コメント2
    ws.row_dimensions[5].height = 93.0   # コメント3
    ws.row_dimensions[6].height = 45.0   # 凡例
    ws.row_dimensions[7].height = 45.0   # ヘッダー
    for d in range(1, days_in_month + 1):
        ws.row_dimensions[7 + d].height = 45.0
    # 空白行と注意書きの行高さ
    ws.row_dimensions[end_row - 1].height = 46.0
    for i in range(len(footer_notes)):
        ws.row_dimensions[end_row + i].height = 49.0

    # ============================================================
    # 印刷設定: A4 縦、1ページに収める
    # ============================================================
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False
    ws.page_margins.left = 0.0
    ws.page_margins.right = 0.0
    ws.page_margins.top = 0.748
    ws.page_margins.bottom = 0.748
    ws.print_area = f"B2:Z{end_row + len(footer_notes) - 1}"

    wb.save(output_path)
    return output_path


# ============================================================
# 動作テスト
# ============================================================

if __name__ == "__main__":
    from .generator import generate_shift, determine_operation_modes
    from .may_2026_data import (
        OFF_REQUESTS, WORK_REQUESTS, PREVIOUS_MONTH_CARRYOVER, FLEXIBLE_OFF_REQUESTS,
    )
    from .rules import MAY_2026_HOLIDAY_OVERRIDES

    print("【シフトExcel出力テスト（テンプレート完全互換・A4縦）】\n")

    print("[1/2] AI でシフトを生成中...")
    modes = determine_operation_modes(2026, 5)
    shift = generate_shift(
        year=2026, month=5,
        off_requests=OFF_REQUESTS, work_requests=WORK_REQUESTS,
        prev_month=PREVIOUS_MONTH_CARRYOVER, flexible_off=FLEXIBLE_OFF_REQUESTS,
        holiday_overrides=MAY_2026_HOLIDAY_OVERRIDES, operation_modes=modes,
        consec_exceptions=["野澤"], max_consec_override=5,
        time_limit_seconds=120, verbose=False,
    )

    if shift is None:
        print("シフト生成失敗")
        exit(1)

    print("[2/2] Excelに書き出し中...")
    output_path = export_shift_to_excel(
        shift=shift,
        output_path=str(OUTPUT_DIR / "2026年5月_AI生成シフト.xlsx"),
        header_comments=[
            "AI 自動生成版です。確認・調整の上ご利用ください。",
            "5月は全体にお休みを増やしています。GW出勤分お体を癒してください。",
            "次回からお休み希望日を「Googleフォーム」に入力していただく予定です。",
        ],
        short_staff_days=[6, 13, 14, 18, 19, 23],
    )
    print(f"  → 保存先: {output_path}")
    print(f"  → ファイルサイズ: {output_path.stat().st_size:,} bytes")
    print("\n✅ 完了。Excelで開いて印刷プレビューを確認してください。")
    print("   ※A4縦1ページに収まる印刷設定です。")
