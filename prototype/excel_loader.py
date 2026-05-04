"""
Excelシフト表読み込み
================================================
data/may_2026_shift.xlsx を MonthlyShift オブジェクトに変換する。

データソース構造:
- 行: 日付（5/1〜5/31）
- 列: 従業員（D=山本〜V=顧問）
- セル: ○/×/△/◆/☆/□/空白
"""

from __future__ import annotations
from datetime import date
from pathlib import Path
from typing import Optional
import warnings

import openpyxl

from .models import MonthlyShift, ShiftAssignment, Store, OperationMode
from .paths import MAY_2026_SHIFT_XLSX


# Excel の列番号と従業員名の対応（5月シフト表）
COLUMN_TO_EMPLOYEE = {
    "D": "山本",
    "E": "板倉",
    "F": "今津",
    "G": "鈴木",
    "H": "田中",
    "I": "岩野",
    "J": "大塚",
    "K": "南",
    "L": "黒澤",
    "M": "牧野",
    "N": "春山",
    "O": "下地",
    "P": "大類",
    "Q": "長尾",
    "R": "野澤",
    "S": "下田",
    "T": "楯",
    "U": "土井",
    "V": "顧問",
}

# 記号 → Store 変換
SYMBOL_TO_STORE = {
    "〇": Store.AKABANE,
    "○": Store.AKABANE,
    "□": Store.HIGASHIGUCHI,
    "△": Store.OMIYA,
    "☆": Store.NISHIGUCHI,
    "◆": Store.SUZURAN,
    "×": Store.OFF,
}


def load_shift_from_excel(
    file_path: str,
    sheet_name: str = "シフト表2026年5月 ",
    year: int = 2026,
    month: int = 5,
) -> tuple[MonthlyShift, list[int]]:
    """
    Excelからシフトを読み込む。

    Returns:
        (MonthlyShift, 人員少マーク日のリスト)
    """
    warnings.filterwarnings("ignore")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    shift = MonthlyShift(year=year, month=month)
    short_staff_days: list[int] = []

    # 行 8〜38 が 5/1〜5/31 のデータ
    for row_num in range(8, 39):
        # 日付を取得
        date_cell = ws.cell(row=row_num, column=2)  # B列
        if date_cell.value is None:
            continue

        # 日付セルから「日」を抽出
        if hasattr(date_cell.value, 'day'):
            day = date_cell.value.day
        else:
            try:
                day = int(str(date_cell.value).split("-")[-1].split()[0])
            except Exception:
                continue

        # 各従業員の配属を読む
        for col_letter, emp_name in COLUMN_TO_EMPLOYEE.items():
            cell = ws[f"{col_letter}{row_num}"]
            value = cell.value
            if value is None:
                # 空白は山本（補助要員）の場合は「投入なし」、それ以外は「未設定」
                # 現状は assignment を作らない（= 休扱いと同等）
                continue
            v = str(value).strip()
            if not v:
                continue
            store = SYMBOL_TO_STORE.get(v)
            if store is None:
                # 未知の記号（例：「○研」など特殊記号）
                # 先頭1文字で判定を試みる
                if v[0] in SYMBOL_TO_STORE:
                    store = SYMBOL_TO_STORE[v[0]]
                else:
                    continue
            shift.assignments.append(ShiftAssignment(
                employee=emp_name, day=day, store=store,
            ))

        # 「人員少」列（Y列）をチェック
        short_cell = ws.cell(row=row_num, column=25)  # Y列
        if short_cell.value is not None and "△" in str(short_cell.value):
            short_staff_days.append(day)

    # 営業モードを設定（5/1-5 は GW REDUCED）
    for day in range(1, 32):
        if 1 <= day <= 5:
            shift.operation_modes[day] = OperationMode.REDUCED
        else:
            shift.operation_modes[day] = OperationMode.NORMAL

    return shift, short_staff_days


if __name__ == "__main__":
    shift, short_days = load_shift_from_excel(
        str(MAY_2026_SHIFT_XLSX)
    )
    print(f"読み込んだシフト: {len(shift.assignments)}件のアサインメント")
    print(f"営業モード設定: {len(shift.operation_modes)}日")
    print(f"人員少マーク日: {short_days}")
    print()
    # 5/1の配置を確認表示
    print("=== 5/1 (金) の配属確認 ===")
    for a in shift.get_day_assignments(1):
        print(f"  {a.employee:6s} → {a.store.display_name}")
    print()
    print("=== 5/3 (日・人員少) の配属確認 ===")
    for a in shift.get_day_assignments(3):
        print(f"  {a.employee:6s} → {a.store.display_name}")
