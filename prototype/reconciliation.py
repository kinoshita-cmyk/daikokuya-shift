"""
提出希望と実績シフトの照合
================================================
過去月のすり合わせ用に、本人提出データと手作業の確定シフトを比較する。

目的:
- 休み希望が実績シフトで勤務になっていないか確認する
- 出勤希望が実績シフトで休み/空白になっていないか確認する
- 過去月のルール調整から除外・注記すべき日を見える化する
"""

from __future__ import annotations

from calendar import monthrange
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import openpyxl

from .models import Store
from .paths import DATA_DIR
from .submission_loader import load_submissions_for_month


DEFAULT_HISTORICAL_WORKBOOK = DATA_DIR / "may_2026_shift.xlsx"

SYMBOL_TO_STORE = {
    "〇": Store.AKABANE,
    "○": Store.AKABANE,
    "□": Store.HIGASHIGUCHI,
    "△": Store.OMIYA,
    "☆": Store.NISHIGUCHI,
    "◆": Store.SUZURAN,
    "×": Store.OFF,
}

STORE_TO_SYMBOL = {
    Store.AKABANE: "○",
    Store.HIGASHIGUCHI: "□",
    Store.OMIYA: "△",
    Store.NISHIGUCHI: "☆",
    Store.SUZURAN: "◆",
    Store.OFF: "×",
}


IGNORED_HISTORICAL_ASSIGNMENTS = {
    ("牧野", Store.HIGASHIGUCHI),
}

NEGOTIATED_HISTORICAL_WORK_OVERRIDES = {
    # 2026年5月は実運用上の口頭調整が確認済み。
    # 「休み希望」は原則絶対だが、過去月のすり合わせでは
    # 初回提出ではなく後出し変更・個別交渉だったものだけを明示的に扱う。
    (2026, 5, "岩野", 1): (Store.AKABANE, "後出し変更後、人員不足のため個別交渉で赤羽勤務"),
    (2026, 5, "下地", 15): (Store.OMIYA, "本人の強い希望による超イレギュラー出勤"),
}


@dataclass
class ReconciliationIssue:
    """提出希望と実績シフトの食い違い1件。"""

    employee: str
    day: int
    issue_type: str
    requested: str
    actual: str
    note: str = ""

    def as_dict(self) -> dict:
        return {
            "氏名": self.employee,
            "日付": f"{self.day}日",
            "種別": self.issue_type,
            "提出希望": self.requested,
            "実績シフト": self.actual or "空白",
            "確認メモ": self.note,
        }


@dataclass
class ReconciledGenerationInputs:
    """過去実績に合わせて補正した生成入力。"""

    off_requests: dict[str, list[int]]
    work_requests: list[tuple[str, int, Optional[Store]]]
    preferred_work_requests: list[tuple[str, int, Optional[Store]]]
    preferred_work_groups: list[tuple[str, list[int], int, Optional[Store]]] = None
    applied_notes: list[str] = None

    @property
    def applied_count(self) -> int:
        return len(self.applied_notes or [])


def normalize_shift_symbol(value) -> str:
    """Excelセル値からシフト記号だけを取り出す。"""
    if value is None:
        return ""
    text = str(value).strip().replace("〇", "○").replace("◯", "○")
    if text in {"|", "｜"}:
        return ""
    for ch in text:
        if ch in SYMBOL_TO_STORE:
            return "○" if ch == "〇" else ch
    return ""


def _store_from_symbol(symbol: str) -> Optional[Store]:
    return SYMBOL_TO_STORE.get(normalize_shift_symbol(symbol))


def _ignore_historical_assignment(employee: str, store: Optional[Store]) -> bool:
    return store is not None and (employee, store) in IGNORED_HISTORICAL_ASSIGNMENTS


def find_historical_sheet_name(workbook, year: int, month: int) -> Optional[str]:
    """対象年月の実績シフトシート名を探す。"""
    candidates = [
        f"シフト表{year}年{month}月",
        f"シフト表{year}年{month}月 ",
        f"シフト表{month}月",
        f"シフト表{month}月 ",
    ]
    for name in candidates:
        if name in workbook.sheetnames:
            return name
    needle = f"{year}年{month}月"
    for name in workbook.sheetnames:
        if needle in name and "シフト表" in name:
            return name
    return None


def find_employee_header_row(ws) -> Optional[int]:
    """山本・板倉などの従業員名が並ぶヘッダー行を探す。"""
    for row_num in range(1, min(ws.max_row, 20) + 1):
        values = [ws.cell(row_num, col).value for col in range(1, ws.max_column + 1)]
        if "山本" in values and "板倉" in values:
            return row_num
    return None


def load_actual_symbols(
    year: int,
    month: int,
    workbook_path: Path = DEFAULT_HISTORICAL_WORKBOOK,
) -> dict[str, dict[int, str]]:
    """実績シフトExcelから {従業員: {日: 記号}} を読み込む。"""
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        return {}

    wb = openpyxl.load_workbook(workbook_path, data_only=False)
    sheet_name = find_historical_sheet_name(wb, year, month)
    if not sheet_name:
        return {}
    ws = wb[sheet_name]
    header_row = find_employee_header_row(ws)
    if header_row is None:
        return {}

    employee_cols: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col).value
        if not isinstance(value, str):
            continue
        name = value.strip()
        if not name or name in {"人員少"}:
            continue
        if name == "専務":
            name = "顧問"
        employee_cols[name] = col

    days_in_month = monthrange(year, month)[1]
    actual: dict[str, dict[int, str]] = {name: {} for name in employee_cols}
    for day in range(1, days_in_month + 1):
        row = header_row + day
        if row > ws.max_row:
            break
        for employee, col in employee_cols.items():
            actual[employee][day] = normalize_shift_symbol(ws.cell(row, col).value)
    return actual


def compare_submissions_to_actual(
    year: int,
    month: int,
    expected_employees: list[str],
    workbook_path: Path = DEFAULT_HISTORICAL_WORKBOOK,
) -> list[ReconciliationIssue]:
    """提出希望と実績シフトを比較し、食い違いを返す。"""
    actual = load_actual_symbols(year, month, workbook_path)
    if not actual:
        return []

    submissions = load_submissions_for_month(year, month, expected_employees)
    issues: list[ReconciliationIssue] = []

    for employee, days in sorted(submissions.off_requests.items()):
        for day in sorted(set(days)):
            actual_symbol = actual.get(employee, {}).get(int(day), "")
            if actual_symbol and actual_symbol != "×":
                issues.append(ReconciliationIssue(
                    employee=employee,
                    day=int(day),
                    issue_type="休み希望なのに勤務",
                    requested="× 休み希望",
                    actual=actual_symbol,
                    note="初回休み希望なら要確認。後出し変更・口頭調整なら個別例外として扱う",
                ))

    work_requests = list(submissions.work_requests)
    work_requests.extend(getattr(submissions, "preferred_work_requests", []))
    seen: set[tuple[str, int, Optional[Store]]] = set()
    for employee, day, requested_store in work_requests:
        key = (employee, int(day), requested_store)
        if key in seen:
            continue
        seen.add(key)
        actual_symbol = actual.get(employee, {}).get(int(day), "")
        requested_label = (
            f"{STORE_TO_SYMBOL.get(requested_store, '')} {requested_store.display_name}"
            if requested_store else "出勤希望"
        )
        if not actual_symbol or actual_symbol == "×":
            issues.append(ReconciliationIssue(
                employee=employee,
                day=int(day),
                issue_type="出勤希望なのに休み/空白",
                requested=requested_label,
                actual=actual_symbol,
                note="出勤希望は希望扱い。調整で休みになることがあります",
            ))
            continue
        if requested_store and actual_symbol != STORE_TO_SYMBOL.get(requested_store):
            issues.append(ReconciliationIssue(
                employee=employee,
                day=int(day),
                issue_type="希望店舗と実績店舗が違う",
                requested=requested_label,
                actual=actual_symbol,
                note="出勤になった場合の希望店舗。調整・合意で別店舗になることがあります",
            ))

    issues.sort(key=lambda x: (x.day, x.employee, x.issue_type))
    return issues


def compare_submissions_to_actual_rows(
    year: int,
    month: int,
    expected_employees: list[str],
    workbook_path: Path = DEFAULT_HISTORICAL_WORKBOOK,
) -> list[dict]:
    """画面表示しやすい dict 形式で照合結果を返す。"""
    return [
        issue.as_dict()
        for issue in compare_submissions_to_actual(
            year, month, expected_employees, workbook_path,
        )
    ]


def build_actual_shift_preferences(
    year: int,
    month: int,
    workbook_path: Path = DEFAULT_HISTORICAL_WORKBOOK,
) -> list[tuple[str, int, Store]]:
    """
    実績シフトをソフト優先条件として返す。

    過去月のすり合わせ用途。ハード固定ではなく、既存の休み希望・人数条件を
    壊さない範囲で実績シフトに寄せるために使う。
    """
    actual = load_actual_symbols(year, month, workbook_path)
    preferences: list[tuple[str, int, Store]] = []
    for employee, day_map in actual.items():
        for day, symbol in day_map.items():
            store = _store_from_symbol(symbol)
            if store is not None:
                if _ignore_historical_assignment(employee, store):
                    continue
                preferences.append((employee, int(day), store))
    return preferences


def reconcile_generation_inputs_with_actual(
    year: int,
    month: int,
    off_requests: dict[str, list[int]],
    work_requests: list[tuple[str, int, Optional[Store]]],
    preferred_work_requests: Optional[list[tuple[str, int, Optional[Store]]]] = None,
    preferred_work_groups: Optional[list[tuple[str, list[int], int, Optional[Store]]]] = None,
    workbook_path: Path = DEFAULT_HISTORICAL_WORKBOOK,
) -> ReconciledGenerationInputs:
    """
    提出希望と実績シフトの差異のうち、明示された口頭調整だけを過去月用に補正する。

    最新運用では、本人の初回「×休み希望」は絶対条件。
    出勤希望・希望店舗は希望扱いなので、実績と違っても自動で休み希望へ変換しない。
    過去月のすり合わせでは、後出し変更や個別交渉が確認済みのものだけを明示補正する。

    将来月では使わず、4月・5月など確定実績とのすり合わせに限定する想定。
    """
    actual = load_actual_symbols(year, month, workbook_path)
    if not actual:
        return ReconciledGenerationInputs(
            off_requests={emp: sorted(set(days)) for emp, days in off_requests.items()},
            work_requests=list(work_requests or []),
            preferred_work_requests=list(preferred_work_requests or []),
            preferred_work_groups=list(preferred_work_groups or []),
            applied_notes=[],
        )

    adjusted_off = {
        emp: set(int(day) for day in days)
        for emp, days in (off_requests or {}).items()
    }
    adjusted_work = list(work_requests or [])
    adjusted_preferred = list(preferred_work_requests or [])
    adjusted_groups = list(preferred_work_groups or [])
    notes: list[str] = []

    def remove_work_for(employee: str, day: int) -> None:
        nonlocal adjusted_work, adjusted_preferred
        adjusted_work = [
            item for item in adjusted_work
            if not (item[0] == employee and int(item[1]) == int(day))
        ]
        adjusted_preferred = [
            item for item in adjusted_preferred
            if not (item[0] == employee and int(item[1]) == int(day))
        ]

    def add_preferred_work(employee: str, day: int, store: Store) -> None:
        remove_work_for(employee, day)
        adjusted_preferred.append((employee, int(day), store))

    # 確認済みの個別交渉だけ、休み希望を外して実績店舗へ寄せる。
    # それ以外の実績差分は、休み希望絶対の原則を壊さないため自動補正しない。
    for key, (store, reason) in NEGOTIATED_HISTORICAL_WORK_OVERRIDES.items():
        override_year, override_month, employee, day = key
        if override_year != year or override_month != month:
            continue
        actual_store = _store_from_symbol(actual.get(employee, {}).get(day, ""))
        if actual_store != store:
            continue
        adjusted_off.setdefault(employee, set()).discard(day)
        add_preferred_work(employee, day, store)
        notes.append(
            f"{employee} {month}/{day}: 個別調整済みとして {STORE_TO_SYMBOL[store]} を優先（{reason}）"
        )

    return ReconciledGenerationInputs(
        off_requests={
            emp: sorted(day for day in days if day)
            for emp, days in adjusted_off.items()
            if days
        },
        work_requests=sorted(
            set(adjusted_work),
            key=lambda item: (item[0], int(item[1]), item[2].name if item[2] else ""),
        ),
        preferred_work_requests=sorted(
            set(adjusted_preferred),
            key=lambda item: (item[0], int(item[1]), item[2].name if item[2] else ""),
        ),
        preferred_work_groups=adjusted_groups,
        applied_notes=list(dict.fromkeys(notes)),
    )
