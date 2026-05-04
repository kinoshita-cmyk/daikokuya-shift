"""
過去のシフトを Validator にかけて検証する
================================================
data/may_2026_shift.xlsx には過去13ヶ月以上のシフトが含まれている。
これらを Validator で検証することで、現状の運用がどの程度ルールを守れているかを確認する。
"""

from __future__ import annotations
import warnings
from datetime import date
from typing import Optional

import openpyxl

from .models import MonthlyShift, ShiftAssignment, Store, OperationMode
from .paths import MAY_2026_SHIFT_XLSX
from .validator import validate
from .excel_loader import COLUMN_TO_EMPLOYEE, SYMBOL_TO_STORE


# 過去シフトのシート名と対応する年月
HISTORICAL_SHEETS = [
    ("シフト表2026年5月 ", 2026, 5),
    ("シフト表2026年4月", 2026, 4),
    ("シフト表2026年3月", 2026, 3),
    ("シフト表2026年2月", 2026, 2),
    ("シフト表2026年1月 ", 2026, 1),
    ("シフト表2025年12月", 2025, 12),
    ("シフト表2025年11月最新", 2025, 11),
    ("シフト表2025年10月", 2025, 10),
    ("シフト表2025年9月", 2025, 9),
    ("シフト表2025年8月", 2025, 8),
    ("シフト表2025年7月", 2025, 7),
    ("シフト表2025年6月", 2025, 6),
]


def find_data_start_row(ws, year: int, month: int) -> Optional[int]:
    """データ開始行を探す。
    複数の行に同じ日付（例: 5/1）が入っているケースがあるので、
    その日付の行のうち、D列以降に記号データがある行を選ぶ。
    """
    candidates = []
    for row_num in range(5, 15):
        cell = ws.cell(row=row_num, column=2)
        if cell.value is None:
            continue
        if hasattr(cell.value, 'day'):
            d = cell.value
            if d.year == year and d.month == month and d.day == 1:
                candidates.append(row_num)

    # 候補のうち、D-V列に記号データがある行
    for row_num in candidates:
        symbol_count = 0
        for col_num in range(4, 23):
            v = ws.cell(row=row_num, column=col_num).value
            if v is not None:
                vs = str(v).strip()
                if vs and (vs[0] in "〇○×△□☆◆" or vs in ("〇", "○", "×", "△", "□", "☆", "◆")):
                    symbol_count += 1
        if symbol_count >= 5:
            return row_num
    return None


def find_employee_columns(ws, header_row: int) -> dict[str, int]:
    """ヘッダー行から従業員名と列番号のマッピングを探す"""
    cols = {}
    for col_num in range(3, 30):
        cell = ws.cell(row=header_row, column=col_num)
        if cell.value is None:
            continue
        v = str(cell.value).strip()
        if v in ("山本", "板倉", "今津", "鈴木", "田中", "岩野", "大塚", "南",
                 "黒澤", "牧野", "春山", "下地", "大類", "長尾", "野澤", "下田",
                 "楯", "土井", "顧問", "専務"):
            # 専務 → 顧問にマッピング
            mapped_name = "顧問" if v == "専務" else v
            cols[mapped_name] = col_num
    return cols


def load_historical_sheet(
    file_path: str, sheet_name: str, year: int, month: int
) -> Optional[MonthlyShift]:
    """過去のシフトシートを読み込む"""
    warnings.filterwarnings("ignore")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]

    # データ開始行を探す
    data_start = find_data_start_row(ws, year, month)
    if data_start is None:
        return None

    # ヘッダー行を探す（data_start以前の行で従業員名が並ぶ行）
    header_row = None
    for hr in range(max(1, data_start - 3), data_start):
        cols = find_employee_columns(ws, hr)
        if len(cols) >= 10:
            header_row = hr
            break

    if header_row is None:
        return None

    emp_cols = find_employee_columns(ws, header_row)
    if len(emp_cols) < 10:
        return None

    from calendar import monthrange
    days_in_month = monthrange(year, month)[1]

    shift = MonthlyShift(year=year, month=month)
    for d in range(1, days_in_month + 1):
        row_num = data_start + (d - 1)
        for emp_name, col_num in emp_cols.items():
            cell = ws.cell(row=row_num, column=col_num)
            if cell.value is None:
                continue
            v = str(cell.value).strip()
            if not v:
                continue
            store = SYMBOL_TO_STORE.get(v)
            if store is None and v[0] in SYMBOL_TO_STORE:
                store = SYMBOL_TO_STORE[v[0]]
            if store is None:
                continue
            shift.assignments.append(ShiftAssignment(
                employee=emp_name, day=d, store=store,
            ))

    # 営業モードの簡易設定（GW期間と他）
    for d in range(1, days_in_month + 1):
        if month == 5 and 1 <= d <= 5:
            shift.operation_modes[d] = OperationMode.REDUCED
        elif month == 8 and 13 <= d <= 16:
            shift.operation_modes[d] = OperationMode.REDUCED
        elif month == 12 and d == 31:
            shift.operation_modes[d] = OperationMode.CLOSED
        elif month == 1 and d in (1, 2):
            shift.operation_modes[d] = OperationMode.CLOSED
        else:
            shift.operation_modes[d] = OperationMode.NORMAL

    return shift


def main():
    print("=" * 70)
    print("【過去シフトの実証検証】")
    print("=" * 70)
    print(f"\n{'月':<10} {'読込件数':<10} {'エラー':<8} {'警告':<8} {'状態'}")
    print("-" * 60)

    file_path = str(MAY_2026_SHIFT_XLSX)

    for sheet_name, year, month in HISTORICAL_SHEETS:
        try:
            shift = load_historical_sheet(file_path, sheet_name, year, month)
            if shift is None or len(shift.assignments) < 50:
                print(f"  {year}年{month:>2}月  読込失敗（シート構造不明）")
                continue
            result = validate(
                shift=shift,
                holiday_overrides={},
                default_holidays=8,
                max_consec=5,
                allow_omiya_short=True,
            )
            status = "✓ OK" if result.error_count == 0 else f"⚠ 違反あり"
            print(f"  {year}年{month:>2}月  "
                  f"{len(shift.assignments):>3}件     "
                  f"{result.error_count:>3}      "
                  f"{result.warning_count:>3}      {status}")
        except Exception as e:
            print(f"  {year}年{month:>2}月  エラー: {type(e).__name__}: {str(e)[:40]}")

    print("\n※ 過去の希望データはないため、希望違反は検出していません")
    print("※ 検証範囲: 連勤(5以下)・店舗人数・大宮アンカー・東口月曜・休日数(8日以上)")


if __name__ == "__main__":
    main()
